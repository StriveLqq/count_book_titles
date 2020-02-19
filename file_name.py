import os
import xlwt
import openpyxl



# def file_list = file_name():
#     file_list = os.listdir("E:\书籍下载")
#     print(file_list)
file_path = 'E:\书库'
file_list1 = []
file_path3 = []
def get_filelist(dir):
    for home, dirs, files in os.walk(dir):
        print("#######dir list#######")
        for dir in dirs:
            print(dir)
            file_path1 = os.path.join(home, dir)
            print(file_path1)
            data_write(file_path1)
        print("#######dir list#######")
        for file in files:  
            file_list1.append(file)                       # 遍历文件
            file_path2 = os.path.join(home, file)   # 获取文件绝对路径  
            file_path3.append(file_path2)            # 将文件路径添加进列表
#  将数据写入新文件
def data_write(file_path):
    file_list = os.listdir(file_path)
    print(file_list)
    f = xlwt.Workbook()
    sheet1 = f.add_sheet(u'sheet1',cell_overwrite_ok=True) #创建sheet
    
    #将数据写入第 i 行，第 j 列
    i = 0
    for data in file_list:
        # for j in range(len(data)):
        sheet1.write(i,0,data)
        i = i + 1
        
    f.save(file_path + 'mulu.xls') #保存文件

def writeExcel(file_list1):
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    i = 1;
    for data in file_list1:
        # for col in range(1,4):
        outws.cell(i, 1).value = data  # 写文件
        i = i+1
    saveExcel = 'E:\书库目录.xlsx'
    outwb.save(saveExcel)  # 一定要记得保存

# file_list = file_name()
# file_path1 = os.path.join(file_path,'\mulu.xls')
# data_write(file_path)
get_filelist(file_path)
writeExcel(file_list1)
# f1 = xlwt.Workbook()
# sheet2 = f1.add_sheet(u'sheet1',cell_overwrite_ok=True) #创建sheet
# i = 0;
# for data in file_list1:
#         # for j in range(len(data)):
#     sheet2.write(i,0,data)
#     i = i + 1

# f1.save('E:\mulu.xls')